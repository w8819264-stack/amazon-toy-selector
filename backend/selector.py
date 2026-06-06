"""
选品引擎 — 根据 config.json 筛选与评分，生成推荐报告
"""
import csv
import json
import os
import math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def load_config(config_path='config.json'):
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def score_product(product, config):
    """
    综合评分模型
    
    权重 (来自 config):
      profit_margin: 35%
      sales_velocity: 25%
      review_count: 15%
      rating: 15%
      competition_index: 10%
    """
    w = config['scoring_weights']
    margin = float(product.get('estimated_profit_margin', 0))
    margin_score = min(100, margin * 200)
    sales = float(product.get('sales_est', 0))
    sales_score = min(100, (sales / 3000) * 100)
    reviews = float(product.get('review_count', 0))
    review_score = min(100, (reviews / 1000) * 100)
    rating = float(product.get('rating', 0))
    rating_score = (rating / 5.0) * 100
    comp_idx = float(product.get('competition_index', 50))
    comp_score = max(0, 100 - comp_idx)
    
    total = (
        margin_score * w['profit_margin'] +
        sales_score * w['sales_velocity'] +
        review_score * w['review_count'] +
        rating_score * w['rating'] +
        comp_score * w['competition_index']
    )
    return round(total, 1), {
        'margin_score': round(margin_score, 1),
        'sales_score': round(sales_score, 1),
        'review_score': round(review_score, 1),
        'rating_score': round(rating_score, 1),
        'comp_score': round(comp_score, 1),
    }


def write_summary_sheet(wb, products, passed, config):
    """写入汇总 Sheet"""
    ws = wb.create_sheet("数据汇总")
    filters = config['filters']
    
    # 用 row/col 直接写，避免解包问题
    rows_data = [
        ("📊 亚马逊玩具选品报告", None),
        ("生成时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("", None),
        ("【筛选条件】", None),
        ("价格区间", f"${filters['price_min']} - ${filters['price_max']}"),
        ("最低评分", str(filters['min_rating'])),
        ("最少评论数", str(filters['min_reviews'])),
        ("最少月销量", str(filters['min_sales_est'])),
        ("最大竞品数", str(filters['max_competitors'])),
        ("最低利润率", f"{filters['min_profit_margin']*100:.0f}%"),
        ("最大重量", f"{filters['weight_max_lbs']} lbs"),
        ("", None),
        ("【筛选结果】", None),
        ("总产品数", str(len(products))),
        ("通过筛选", str(len(passed))),
        ("强烈推荐", str(sum(1 for p in passed if '强烈推荐' in str(p.get('recommendation', ''))))),
        ("推荐", str(sum(1 for p in passed if str(p.get('recommendation', '')) == '👍 推荐'))),
        ("可考虑", str(sum(1 for p in passed if '可考虑' in str(p.get('recommendation', ''))))),
    ]
    
    # 动态统计行
    if passed:
        avg_price = sum(float(p['price']) for p in passed) / len(passed)
        avg_margin = sum(float(p['estimated_profit_margin']) for p in passed) / len(passed)
        avg_profit = sum(float(p['net_profit']) for p in passed) / len(passed)
        rows_data.append(("", None))
        rows_data.append(("【平均数据 (通过筛选的产品)】", None))
        rows_data.append(("平均售价", f"${avg_price:.2f}"))
        rows_data.append(("平均利润率", f"{avg_margin*100:.1f}%"))
        rows_data.append(("平均净利润", f"${avg_profit:.2f}"))
    
    for r, (label, value) in enumerate(rows_data, 1):
        cell_a = ws.cell(row=r, column=1, value=label)
        if value is None:
            cell_a.font = Font(bold=True, size=12, color='2B579A')
        else:
            cell_a.font = Font(bold=True, size=11)
            ws.cell(row=r, column=2, value=value).font = Font(size=11)
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 32


def select_products(input_csv='data/processed_toys.csv', output_excel='data/toy_selection.xlsx', config_path='config.json'):
    """
    选品主流程: 读取数据 → 筛选 → 评分 → 生成 Excel 报告
    """
    config = load_config(config_path)
    filters = config['filters']
    
    # 读取数据
    products = []
    with open(input_csv, 'r', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            products.append(row)
    
    print(f"📊 待筛选产品: {len(products)} 个")
    
    # ── 硬性筛选 ──
    passed = []
    rejected_reasons = {}
    for p in products:
        price = float(p['price'])
        rating = float(p['rating'])
        reviews = int(p['review_count'])
        sales = int(p['sales_est'])
        comp = int(p['competitors'])
        margin = float(p['estimated_profit_margin'])
        weight = float(p['weight_lbs'])
        l, w, h = float(p['length_in']), float(p['width_in']), float(p['height_in'])
        
        reason = None
        if price < filters['price_min']:
            reason = f"价格 ${price:.2f}"
        elif price > filters['price_max']:
            reason = f"价格 ${price:.2f}"
        elif rating < filters['min_rating']:
            reason = f"评分 {rating}"
        elif reviews < filters['min_reviews']:
            reason = f"评论 {reviews}"
        elif sales < filters['min_sales_est']:
            reason = f"销量 {sales}"
        elif comp > filters['max_competitors']:
            reason = f"竞品 {comp}"
        elif margin < filters['min_profit_margin']:
            reason = f"利润率 {(margin*100):.0f}%"
        elif weight > filters['weight_max_lbs']:
            reason = f"重量 {weight}lbs"
        elif (l > filters['dimensions_max']['length_in'] or 
              w > filters['dimensions_max']['width_in'] or 
              h > filters['dimensions_max']['height_in']):
            reason = f"尺寸 {l}x{w}x{h}"
        
        if reason:
            rejected_reasons[reason] = rejected_reasons.get(reason, 0) + 1
        else:
            total_score, detail = score_product(p, config)
            p['total_score'] = total_score
            for k, v in detail.items():
                p[k] = v
            if total_score >= 75:
                p['recommendation'] = '⭐ 强烈推荐'
            elif total_score >= 55:
                p['recommendation'] = '👍 推荐'
            else:
                p['recommendation'] = '👀 可考虑'
            passed.append(p)
    
    # 排序
    passed.sort(key=lambda x: x['total_score'], reverse=True)
    
    print(f"✅ 通过筛选: {len(passed)} 个")
    print(f"❌ 被淘汰: {len(products) - len(passed)} 个")
    for reason, count in sorted(rejected_reasons.items(), key=lambda x: -x[1])[:10]:
        print(f"   - {reason}: {count}个")
    
    # ── 生成 Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "玩具选品推荐"
    
    # 样式
    header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    light_blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['排名', '推荐等级', '综合评分', 'ASIN', '产品标题', '子类目', '售价($)', 
               '评分', '评论数', '月销量', '竞品数', '利润率', '净利润($)', 
               'FBA费用($)', '重量(lbs)', '利润评分', '销量评分', '评论评分', '评分评分', '竞争评分']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # 数据行
    for idx, p in enumerate(passed):
        row = idx + 2
        values = [
            idx + 1,
            p['recommendation'],
            float(p['total_score']),
            p['asin'],
            p['title'],
            p['subcategory'],
            float(p['price']),
            float(p['rating']),
            int(p['review_count']),
            int(p['sales_est']),
            int(p['competitors']),
            f"{float(p['estimated_profit_margin'])*100:.1f}%",
            float(p['net_profit']),
            float(p['total_fba_cost']),
            float(p['weight_lbs']),
            float(p.get('margin_score', 0)),
            float(p.get('sales_score', 0)),
            float(p.get('review_score', 0)),
            float(p.get('rating_score', 0)),
            float(p.get('comp_score', 0)),
        ]
        
        if '强烈推荐' in p['recommendation']:
            row_fill = green_fill
        elif '推荐' in p['recommendation']:
            row_fill = light_blue_fill
        else:
            row_fill = yellow_fill
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.fill = row_fill
            if col == 4:
                cell.font = Font(name='Consolas', size=10)
    
    # 列宽 & 冻结
    col_widths = [6, 14, 10, 14, 45, 20, 10, 7, 10, 10, 8, 9, 11, 11, 9, 10, 10, 10, 10, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(passed)+1}"
    
    # 汇总 Sheet
    write_summary_sheet(wb, products, passed, config)
    
    # 保存
    os.makedirs(os.path.dirname(output_excel), exist_ok=True)
    wb.save(output_excel)
    print(f"\n📁 选品报告已生成: {output_excel}")
    print(f"   ├─ Sheet1: 玩具选品推荐 ({len(passed)} 条)")
    print(f"   └─ Sheet2: 数据汇总")
    
    return passed


if __name__ == '__main__':
    if not os.path.exists('data/raw_toys.csv'):
        from collector import collect_toy_data, save_raw_data
        save_raw_data(collect_toy_data(150))
    if not os.path.exists('data/processed_toys.csv'):
        from processor import process_products
        process_products()
    select_products()
